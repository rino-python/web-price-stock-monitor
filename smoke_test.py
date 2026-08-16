from io import BytesIO

from openpyxl import load_workbook

from generate_sample import DEMO_HTML, SAMPLES_DIR, previous_dataframe
from src.excel_io import build_result_xlsx
from src.run import run_watch


def main() -> None:
    result = run_watch(
        sample_html=DEMO_HTML,
        previous=previous_dataframe(),
        previous_name="先週取得.csv",
    )
    assert result.item_count == 6
    assert result.issue_count >= 1
    assert result.change_count >= 3
    xlsx = build_result_xlsx(result)
    wb = load_workbook(BytesIO(xlsx))
    assert wb.sheetnames == ["今回一覧", "変化", "失敗ログ", "取得ルール", "処理メモ"]
    out = SAMPLES_DIR / "ウォッチ結果_確認用.xlsx"
    SAMPLES_DIR.mkdir(exist_ok=True)
    out.write_bytes(xlsx)
    print("ok", result.item_count, "items", result.change_count, "changes", len(xlsx), "bytes")


if __name__ == "__main__":
    main()
