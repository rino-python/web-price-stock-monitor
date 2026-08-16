"""取得結果を、人が開いてすぐ使える Excel バイト列にする。"""

from __future__ import annotations

import math
import unicodedata
from io import BytesIO

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook

from src.colors import CHANGE_HEX
from src.run import WatchResult
from src.shop_config import MAX_ITEMS, MAX_PAGES

HEADER_FILL = PatternFill("solid", fgColor="0D7377")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
CHANGE_FILLS = {label: PatternFill("solid", fgColor=hex_) for label, hex_ in CHANGE_HEX.items()}


def _style_header(ws, max_col: int | None = None, max_row: int | None = None) -> None:
    last_col = max_col or ws.max_column
    last_row = max_row or ws.max_row
    for col in range(1, last_col + 1):
        cell = ws.cell(1, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"


def _display_width(text: str) -> float:
    width = 0.0
    for char in text:
        width += 2.0 if unicodedata.east_asian_width(char) in {"W", "F"} else 1.0
    return width


def _is_number(value: object) -> bool:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return False
    return not math.isnan(float(value))


def _cell_display_text(cell) -> str:
    value = cell.value
    if value is None:
        return ""
    fmt = cell.number_format or "General"
    if _is_number(value):
        if ".00" in fmt:
            return f"{value:,.2f}"
        if "¥" in fmt or "£" in fmt or "#,##0" in fmt:
            return f"{int(round(float(value))):,}"
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value)


def _autosize(ws, min_width: float = 8, max_width: float = 40, header_extra: float = 3.0) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        longest = 0.0
        for index, cell in enumerate(col):
            width = _display_width(_cell_display_text(cell))
            if index == 0:
                width += header_extra
            longest = max(longest, width)
        ws.column_dimensions[letter].width = min(max_width, max(min_width, longest + 1.2))


def _write_df(ws, df) -> None:
    if df.empty:
        ws.append(list(df.columns) if len(df.columns) else ["（行がありません）"])
        return
    for row in dataframe_to_rows(df, index=False, header=True):
        cleaned = []
        for value in row:
            if isinstance(value, float) and math.isnan(value):
                cleaned.append(None)
            else:
                cleaned.append(value)
        ws.append(cleaned)


def _border_rows(ws) -> None:
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = THIN


def build_result_xlsx(result: WatchResult) -> bytes:
    wb = Workbook()

    ws_list = wb.active
    ws_list.title = "今回一覧"
    listing = result.listing.copy()
    _write_df(ws_list, listing)
    _style_header(ws_list)
    for row in ws_list.iter_rows(min_row=2, max_row=ws_list.max_row):
        for cell in row:
            cell.border = THIN
            name = ws_list.cell(1, cell.column).value
            if name == "価格" and _is_number(cell.value):
                cell.number_format = "¥#,##0"
    if ws_list.max_row == 1:
        ws_list.append(["（取得できた商品はありません）"])
    _autosize(ws_list)

    ws_chg = wb.create_sheet("変化")
    _write_df(ws_chg, result.changes)
    _style_header(ws_chg)
    headers = [cell.value for cell in ws_chg[1]]
    label_idx = headers.index("変化") + 1 if "変化" in headers else None
    for row in ws_chg.iter_rows(min_row=2, max_row=ws_chg.max_row):
        fill = None
        if label_idx:
            fill = CHANGE_FILLS.get(str(row[label_idx - 1].value or ""))
        for cell in row:
            cell.border = THIN
            if fill is not None:
                cell.fill = fill
            name = ws_chg.cell(1, cell.column).value
            if name in {"今回価格", "前回価格"} and _is_number(cell.value):
                cell.number_format = "¥#,##0"
    if ws_chg.max_row == 1:
        ws_chg.append(["（前回と比べた変化はありません）"])
    _autosize(ws_chg)

    ws_fail = wb.create_sheet("失敗ログ")
    _write_df(ws_fail, result.issues)
    _style_header(ws_fail)
    _border_rows(ws_fail)
    if ws_fail.max_row == 1:
        ws_fail.append(["（失敗はありません）"])
    _autosize(ws_fail)

    ws_rule = wb.create_sheet("取得ルール")
    rules = [
        ("項目", "内容"),
        ("対象", "自社の公開カタログのみ。デモは架空店「和洋雑貨 かもめ堂」。他社サイトや実在店舗は叩きません。"),
        ("納品時", "src/shop_config.py の URL とセレクタを御社の公開一覧に差し替えます。"),
        ("robots.txt", "取得前に確認します。禁止されていれば止めます。"),
        ("間隔", "ページとページのあいだは 1.2 秒空けます。"),
        ("リトライ", "失敗したら最大 3 回までやり直します。"),
        ("上限", f"1回あたり最大 {MAX_PAGES} ページ、{MAX_ITEMS} 件。引数ではこれ以上にできません。"),
        ("取らないもの", "他社URL、ログインが必要なページ、個人情報、大手モール。"),
        ("通知", "Slack やメールは送りません。変化シートと画面の件数で知らせます。"),
        ("Sheets", "デモは Excel です。案件ではスプレッドシート追記に差し替えます。"),
    ]
    for row in rules:
        ws_rule.append(list(row))
    _style_header(ws_rule)
    _border_rows(ws_rule)
    wrap = Alignment(wrap_text=True, vertical="center")
    for row in ws_rule.iter_rows(min_row=2, min_col=2, max_col=2, max_row=ws_rule.max_row):
        for cell in row:
            cell.alignment = wrap
            ws_rule.row_dimensions[cell.row].height = 32
    ws_rule.column_dimensions["A"].width = 16
    ws_rule.column_dimensions["B"].width = 72

    ws_meta = wb.create_sheet("処理メモ")
    meta_rows = [
        ("項目", "内容"),
        ("取得元", result.source_url),
        ("前回ファイル", result.previous_name),
        ("取得ページ数", result.page_count),
        ("取得件数", result.item_count),
        ("件数上限", MAX_ITEMS),
        ("ページ上限", MAX_PAGES),
        ("変化件数", result.change_count),
        ("失敗件数", result.issue_count),
        ("サンプル使用", "はい" if result.used_sample else "いいえ"),
        ("robots.txt", result.robots_note),
    ]
    for row in meta_rows:
        ws_meta.append(list(row))
    _style_header(ws_meta)
    left = Alignment(horizontal="left", vertical="center")
    wrap_b = Alignment(wrap_text=True, horizontal="left", vertical="center")
    for row in ws_meta.iter_rows(min_row=2, min_col=2, max_col=2, max_row=ws_meta.max_row):
        for cell in row:
            cell.alignment = wrap_b if cell.row == ws_meta.max_row else left
            if cell.row == ws_meta.max_row:
                ws_meta.row_dimensions[cell.row].height = 36
    _autosize(ws_meta)
    ws_meta.column_dimensions["B"].width = 72

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
